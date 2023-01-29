import matplotlib.pyplot as plt
import geopandas
import pandas
import openpyxl as xl

wb = xl.load_workbook('States_shapefile.xlsx')

sheet1 = wb['States_shapefile']
listofTZs = []


for x in range(2,sheet1.max_row+1):
    listofTZs.append(sheet1.cell(x,2).value)
print(listofTZs)

US =  geopandas.read_file('States_shapefile.zip')
US = US.assign(TZ=listofTZs)
print(US.columns.tolist())

print(US.head())

fuck=US.plot(column='TZ',  cmap='OrRd')
plt.show()

plt.savefig('fuck.jpg')



"""
world = geopandas.read_file(geopandas.datasets.get_path('naturalearth_lowres'))

world = world[(world.pop_est>0) & (world.name!="Antarctica")]
world['gdp_per_cap'] = world.gdp_md_est / world.pop_est
world.plot(column='gdp_per_cap',legend=True)

plt.savefig('fuck.jpg')
"""